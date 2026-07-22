import openpyxl
import sys

file_path = "Time-Table, FSC, Fall-2025.xlsx"
try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    print(f"Active Sheet: {sheet.title}")
    
    # Check first 5 rows
    for row in range(1, 6):
        print(f"--- Row {row} ---")
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            if cell.value is not None:
                fill_color = None
                if cell.fill and cell.fill.start_color and cell.fill.start_color.index:
                    fill_color = cell.fill.start_color.index
                    if hasattr(fill_color, 'rgb'):
                        fill_color = fill_color.rgb
                print(f"Cell({row},{col}) Value: {cell.value} | Color: {fill_color}")
except Exception as e:
    print(f"Error: {e}")
